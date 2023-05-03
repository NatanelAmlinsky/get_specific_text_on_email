import openpyxl
import win32com.client


class OutlookAccount:
    def __init__(self, account_name):
        self.outlook = win32com.client.Dispatch("Outlook.Application")
        self.namespace = self.outlook.GetNamespace("MAPI")
        self.account_name = account_name
        self.account = None

    def login(self):
        for a in self.namespace.Accounts:
            if a.DisplayName == self.account_name:
                self.account = a
                break

        if not self.account:
            print(f"Could not find account '{self.account_name}'")
            return False

        self.inbox_folder = self.account.DeliveryStore.GetDefaultFolder(6)
        return True

    def get_email_content(message):
        email_body = message.body
        parts = []

        for line in email_body.splitlines():
            parts.extend(line.split("\n"))

        order_info = {}

        for element in parts:
            if "מס' הזמנה:" in element:
                order_info["Order Number"] = element.split(": ")[1]
                order_number = order_info["Order Number"]
            elif "שם מלא:" in element:
                order_info["Full Name"] = element.split(": ")[1]
                full_name = order_info["Full Name"]
            elif "כתובת:" in element:
                order_info["Address"] = element.split(": ")[1]
                address = order_info["Address"]
            elif "אימייל:" in element:
                order_info["Email"] = element.split(": ")[1]
                email = order_info["Email"]
            elif "מס' ליצירת קשר:" in element:
                order_info["Phone Number"] = element.split(": ")[1]
                phone_number = order_info["Phone Number"]
            elif "הספרים שנבחרו:" in element:
                order_info["Books"] = element.split(": ")[1]
                books = order_info["Books"]
            elif "IP:" in element:
                order_info["IP Address"] = element.split(": ")[1]
                ip_address = order_info["IP Address"]

        # Create a new workbook and select the active worksheet
        wb = openpyxl.load_workbook("C:\\Users\\natan\\Desktop\\EmailAutomation\\shipping_info.xlsx")
        ws = wb.active

        # Write the headers to the first row of the worksheet if the worksheet is empty
        if not any(ws.iter_rows()):
            headers = ["Order Number", "Full Name", "Address", "Email", "Phone Number", "Books", "IP Address"]
            ws.append(headers)

        # Find the first empty row
        current_row = 2
        while ws.cell(row=current_row, column=1).value is not None:
            current_row += 1

        # Write the order info to the empty row
        row = [order_info["Order Number"], order_info["Full Name"], order_info["Address"], order_info["Email"],
               order_info["Phone Number"], order_info["Books"], order_info["IP Address"]]
        for i, value in enumerate(row):
            ws.cell(row=current_row, column=i + 1, value=value)

        # Save the workbook to a file
        wb.save("C:\\Users\\natan\\Desktop\\EmailAutomation\\shipping_info.xlsx")
        return print(order_info)





